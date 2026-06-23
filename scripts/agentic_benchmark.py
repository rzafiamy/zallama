#!/usr/bin/env python3
"""
agentic_benchmark.py — Benchmark Zallama text models on agentic capability.

Measures three axes that matter for agent use:
  1. tool_calling   — picks the right tool and emits valid JSON arguments
  2. reasoning      — multi-step / chained reasoning to a checkable answer
  3. instruction    — strict instruction & output-format following

Runs models SEQUENTIALLY (unload-before-load) because this box has ~7.8 GB
VRAM and only one text model fits at a time. Scoring is deterministic
(temperature=0) and programmatic; full raw responses are written to Excel for
manual review.

Usage:
    python3 scripts/agentic_benchmark.py
    python3 scripts/agentic_benchmark.py --models qwen3.5-9b-q4_k_m gemma-4-e4b-it-q4_k_m
    python3 scripts/agentic_benchmark.py --base-url http://localhost:11435 --out bench.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import time
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_MODELS = [
    "qwen3.5-4b-q4_k_m",
    "qwen3.5-4b-ud-q4_k_xl",
    "qwen3.5-9b-q4_k_m",
    "gemma-4-e4b-it-q4_k_m",
]

# ---------------------------------------------------------------------------
# Tool definitions shared by the tool-calling tasks
# ---------------------------------------------------------------------------
TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "get_weather",
            "description": "Get the current weather for a given city.",
            "parameters": {
                "type": "object",
                "properties": {
                    "city": {"type": "string", "description": "City name"},
                    "unit": {"type": "string", "enum": ["celsius", "fahrenheit"]},
                },
                "required": ["city"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "send_email",
            "description": "Send an email to a recipient.",
            "parameters": {
                "type": "object",
                "properties": {
                    "to": {"type": "string"},
                    "subject": {"type": "string"},
                    "body": {"type": "string"},
                },
                "required": ["to", "subject", "body"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "calculate",
            "description": "Evaluate an arithmetic expression and return the number.",
            "parameters": {
                "type": "object",
                "properties": {"expression": {"type": "string"}},
                "required": ["expression"],
            },
        },
    },
]


# ---------------------------------------------------------------------------
# Scorers: each returns (passed: bool, note: str)
# ---------------------------------------------------------------------------
def score_tool_call(resp_msg, want_name, want_args_check):
    """Verify the model emitted a tool_call with the right name and args."""
    calls = (resp_msg or {}).get("tool_calls") or []
    if not calls:
        return False, "no tool_call emitted"
    fn = calls[0].get("function", {})
    name = fn.get("name")
    if name != want_name:
        return False, f"called '{name}', expected '{want_name}'"
    try:
        args = json.loads(fn.get("arguments") or "{}")
    except json.JSONDecodeError:
        return False, f"args not valid JSON: {fn.get('arguments')!r}"
    ok, note = want_args_check(args)
    return ok, note or f"tool '{name}' ok"


def score_no_tool(resp_msg, _content):
    """Pass when the model correctly answers WITHOUT calling a tool."""
    calls = (resp_msg or {}).get("tool_calls") or []
    if calls:
        return False, f"called tool '{calls[0].get('function', {}).get('name')}' when none needed"
    return True, "no tool called (correct)"


def contains_number(content, target, tol=1e-6):
    nums = re.findall(r"-?\d+(?:\.\d+)?", content.replace(",", ""))
    for n in nums:
        try:
            if abs(float(n) - target) <= tol:
                return True
        except ValueError:
            pass
    return False


def score_exact_json(content, expected):
    """Pass when content parses to exactly the expected JSON object."""
    txt = content.strip()
    m = re.search(r"\{.*\}", txt, re.DOTALL)
    if not m:
        return False, "no JSON object found"
    try:
        got = json.loads(m.group(0))
    except json.JSONDecodeError as e:
        return False, f"JSON parse error: {e}"
    if got == expected:
        return True, "exact JSON match"
    return False, f"got {json.dumps(got)} vs {json.dumps(expected)}"


# ---------------------------------------------------------------------------
# Task suite
# ---------------------------------------------------------------------------
def build_tasks():
    tasks = []

    # ---- TOOL CALLING ----
    tasks.append(dict(
        cat="tool_calling", id="weather_basic",
        messages=[{"role": "user", "content": "What's the weather in Tokyo right now?"}],
        tools=TOOLS,
        score=lambda m, c: score_tool_call(
            m, "get_weather",
            lambda a: (a.get("city", "").lower() == "tokyo", f"city={a.get('city')!r}")),
    ))
    tasks.append(dict(
        cat="tool_calling", id="email_fields",
        messages=[{"role": "user", "content":
                   "Email bob@acme.com with the subject 'Lunch' and body 'See you at noon'."}],
        tools=TOOLS,
        score=lambda m, c: score_tool_call(
            m, "send_email",
            lambda a: (a.get("to") == "bob@acme.com" and "lunch" in (a.get("subject", "").lower())
                       and bool(a.get("body")),
                       f"to={a.get('to')!r} subj={a.get('subject')!r}")),
    ))
    tasks.append(dict(
        cat="tool_calling", id="calc_tool",
        messages=[{"role": "user", "content": "Use a tool to compute 17 * 23 + 9."}],
        tools=TOOLS,
        score=lambda m, c: score_tool_call(
            m, "calculate",
            lambda a: ("17" in a.get("expression", "") and "23" in a.get("expression", ""),
                       f"expr={a.get('expression')!r}")),
    ))
    tasks.append(dict(
        cat="tool_calling", id="no_tool_needed",
        messages=[{"role": "user", "content":
                   "Say hello and tell me a fun fact. Do NOT use any tools."}],
        tools=TOOLS,
        score=score_no_tool,
    ))

    # ---- MULTI-STEP REASONING ----
    tasks.append(dict(
        cat="reasoning", id="word_problem",
        messages=[{"role": "user", "content":
                   "A train leaves at 2:00 PM going 60 mph. Another leaves the same station at "
                   "3:00 PM going 90 mph in the same direction. At what time (in hours past 2 PM) "
                   "does the second train catch the first? Give the final number of hours past 2 PM."}],
        score=lambda m, c: (contains_number(c, 3.0),
                            "expects 3 (caught at 5PM = 3h past 2PM)"),
    ))
    tasks.append(dict(
        cat="reasoning", id="multi_step_arith",
        messages=[{"role": "user", "content":
                   "Start with 100. Subtract 35. Multiply the result by 2. Then add 15. "
                   "Finally divide by 5. What is the final number?"}],
        score=lambda m, c: (contains_number(c, 29.0), "expects 29"),
    ))
    tasks.append(dict(
        cat="reasoning", id="logic_deduction",
        messages=[{"role": "user", "content":
                   "Alice is older than Bob. Carol is younger than Bob. Dave is older than Alice. "
                   "List all four from youngest to oldest as a comma-separated line."}],
        score=lambda m, c: (
            bool(re.search(r"carol.*bob.*alice.*dave", c.lower(), re.DOTALL)),
            "expects order Carol, Bob, Alice, Dave"),
    ))
    tasks.append(dict(
        cat="reasoning", id="planning",
        messages=[{"role": "user", "content":
                   "You have 3 tasks: (A) takes 2h, (B) takes 1h and must finish before A starts, "
                   "(C) takes 3h and can run anytime. Working strictly one task at a time, what is "
                   "the minimum total hours to finish all three? Give just the number."}],
        score=lambda m, c: (contains_number(c, 6.0), "expects 6 (2+1+3)"),
    ))

    # ---- INSTRUCTION / FORMAT FOLLOWING ----
    tasks.append(dict(
        cat="instruction", id="json_only",
        messages=[{"role": "user", "content":
                   'Respond with ONLY this JSON, no prose: {"status":"ok","count":3}'}],
        score=lambda m, c: score_exact_json(c, {"status": "ok", "count": 3}),
    ))
    tasks.append(dict(
        cat="instruction", id="extract_json",
        messages=[{"role": "user", "content":
                   "From 'John Smith, age 42, lives in Berlin' produce a JSON object with keys "
                   'name, age, city. Output only the JSON.'}],
        score=lambda m, c: score_exact_json(
            c, {"name": "John Smith", "age": 42, "city": "Berlin"}),
    ))
    tasks.append(dict(
        cat="instruction", id="word_limit",
        messages=[{"role": "user", "content":
                   "Describe the ocean in EXACTLY three words. Output only those three words."}],
        score=lambda m, c: (len([w for w in re.findall(r"[A-Za-z']+", c)]) == 3,
                            "expects exactly 3 words"),
    ))
    tasks.append(dict(
        cat="instruction", id="format_list",
        messages=[{"role": "user", "content":
                   "List three primary colors, each on its own line, each prefixed with '- '. "
                   "No other text."}],
        score=lambda m, c: (
            len([ln for ln in c.strip().splitlines() if ln.strip().startswith("- ")]) == 3,
            "expects exactly 3 lines each starting with '- '"),
    ))

    return tasks


# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------
def unload_all_text_models(base, models, timeout=60):
    for name in models:
        try:
            requests.post(f"{base}/api/models/{name}/unload", timeout=timeout)
        except requests.RequestException:
            pass


def run_task(base, model, task, timeout=180):
    payload = {
        "model": model,
        "messages": task["messages"],
        "temperature": 0,
        # Generous budget so reasoning models (e.g. the 9B) don't hit the
        # length cap inside their think phase before emitting the answer.
        "max_tokens": 2048,
    }
    if task.get("tools"):
        payload["tools"] = task["tools"]
        payload["tool_choice"] = "auto"

    t0 = time.time()
    try:
        r = requests.post(f"{base}/v1/chat/completions", json=payload, timeout=timeout)
    except requests.RequestException as e:
        return dict(passed=False, note=f"request error: {e}", content="", raw="",
                    latency=time.time() - t0, tok=0, tps=0.0)
    latency = time.time() - t0

    if r.status_code != 200:
        return dict(passed=False, note=f"HTTP {r.status_code}: {r.text[:200]}",
                    content="", raw=r.text[:500], latency=latency, tok=0, tps=0.0)

    data = r.json()
    msg = data["choices"][0]["message"]
    content = msg.get("content") or ""
    # Reasoning models may leave content empty and put the final answer at the
    # tail of reasoning_content. Fall back to it so they're judged on the same
    # visible output as non-reasoning models.
    reasoning = msg.get("reasoning_content") or ""
    if not content.strip() and reasoning.strip():
        content = reasoning
    usage = data.get("usage", {})
    out_tok = usage.get("completion_tokens", 0)
    tps = (out_tok / latency) if latency > 0 and out_tok else 0.0

    passed, note = task["score"](msg, content)
    raw = json.dumps(msg, ensure_ascii=False)[:1500]
    return dict(passed=passed, note=note, content=content[:1500], raw=raw,
                latency=latency, tok=out_tok, tps=tps)


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------
HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="305496")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
WRAP = Alignment(wrap_text=True, vertical="top")


def write_excel(path, models, tasks, results, meta):
    wb = Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    cats = ["tool_calling", "reasoning", "instruction"]
    header = ["Model", "Overall %"] + [f"{c} %" for c in cats] + \
             ["Avg latency (s)", "Avg tok/s"]
    ws.append(header)
    for col, _ in enumerate(header, 1):
        cell = ws.cell(1, col)
        cell.font = HDR
        cell.fill = HDR_FILL

    for model in models:
        rows = [r for r in results if r["model"] == model]
        if not rows:
            ws.append([model, "ERROR"]); continue
        overall = 100 * sum(r["passed"] for r in rows) / len(rows)
        cat_scores = []
        for c in cats:
            cr = [r for r in rows if r["cat"] == c]
            cat_scores.append(round(100 * sum(x["passed"] for x in cr) / len(cr), 1) if cr else 0)
        lat = sum(r["latency"] for r in rows) / len(rows)
        tps_vals = [r["tps"] for r in rows if r["tps"] > 0]
        avg_tps = sum(tps_vals) / len(tps_vals) if tps_vals else 0
        ws.append([model, round(overall, 1)] + cat_scores +
                  [round(lat, 2), round(avg_tps, 1)])

    # rank highlight: bold the best Overall %
    best_row, best_val = None, -1
    for ridx in range(2, ws.max_row + 1):
        v = ws.cell(ridx, 2).value
        if isinstance(v, (int, float)) and v > best_val:
            best_val, best_row = v, ridx
    if best_row:
        for col in range(1, len(header) + 1):
            ws.cell(best_row, col).font = Font(bold=True)

    widths = [26, 11, 16, 14, 16, 16, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Details sheet ----
    wd = wb.create_sheet("Details")
    dh = ["Model", "Category", "Task", "Pass", "Score note", "Latency (s)",
          "Out tok", "tok/s", "Prompt", "Response (truncated)"]
    wd.append(dh)
    for col, _ in enumerate(dh, 1):
        wd.cell(1, col).font = HDR
        wd.cell(1, col).fill = HDR_FILL

    task_prompt = {t["id"]: t["messages"][-1]["content"] for t in tasks}
    for r in results:
        wd.append([
            r["model"], r["cat"], r["id"], "PASS" if r["passed"] else "FAIL",
            r["note"], round(r["latency"], 2), r["tok"], round(r["tps"], 1),
            task_prompt.get(r["id"], ""),
            r["content"] or r["raw"],
        ])
        fill = PASS_FILL if r["passed"] else FAIL_FILL
        wd.cell(wd.max_row, 4).fill = fill

    for col, w in zip(range(1, len(dh) + 1), [22, 14, 18, 7, 34, 11, 9, 8, 50, 70]):
        wd.column_dimensions[get_column_letter(col)].width = w
    for ridx in range(2, wd.max_row + 1):
        for col in (5, 9, 10):
            wd.cell(ridx, col).alignment = WRAP

    # ---- Meta sheet ----
    wm = wb.create_sheet("Run info")
    for k, v in meta.items():
        wm.append([k, str(v)])
    wm.column_dimensions["A"].width = 22
    wm.column_dimensions["B"].width = 60

    wb.save(path)


# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--base-url", default="http://localhost:11435")
    ap.add_argument("--models", nargs="*", default=DEFAULT_MODELS)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    base = args.base_url.rstrip("/")
    out = args.out or f"agentic_benchmark_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    tasks = build_tasks()

    print(f"Benchmark: {len(args.models)} models x {len(tasks)} tasks -> {out}\n")
    results = []
    start = time.time()

    for model in args.models:
        print(f"=== {model} ===")
        # Ensure only this text model is resident: unload every other text model.
        unload_all_text_models(base, [m for m in args.models if m != model])
        time.sleep(1)
        for task in tasks:
            res = run_task(base, model, task)
            res.update(model=model, cat=task["cat"], id=task["id"])
            results.append(res)
            flag = "PASS" if res["passed"] else "FAIL"
            print(f"  [{flag}] {task['cat']:<12} {task['id']:<18} "
                  f"{res['latency']:.1f}s {res['tps']:.0f} tok/s — {res['note'][:60]}")
        # free VRAM for the next model
        unload_all_text_models(base, [model])
        print()

    meta = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "base_url": base,
        "models": ", ".join(args.models),
        "num_tasks": len(tasks),
        "total_runtime_s": round(time.time() - start, 1),
        "scoring": "deterministic (temperature=0), programmatic pass/fail",
        "note": "One text model resident at a time (sequential, unload-before-load) due to ~7.8GB VRAM.",
    }
    write_excel(out, args.models, tasks, results, meta)
    print(f"Wrote {out}")

    # quick console leaderboard
    print("\nLeaderboard (overall pass %):")
    board = []
    for m in args.models:
        rows = [r for r in results if r["model"] == m]
        if rows:
            board.append((m, 100 * sum(r["passed"] for r in rows) / len(rows)))
    for m, pct in sorted(board, key=lambda x: -x[1]):
        print(f"  {pct:5.1f}%  {m}")


if __name__ == "__main__":
    main()
